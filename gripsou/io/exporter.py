"""Export budget data to xlsx or csv."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from gripsou.db.models import LineItem

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

_HEADER_FILL = PatternFill("solid", fgColor="1565C0")
_CREDIT_FILL = PatternFill("solid", fgColor="E8F5E9")
_DEBIT_FILL = PatternFill("solid", fgColor="FFEBEE")
_SUMMARY_FILL = PatternFill("solid", fgColor="E3F2FD")
_POS_FONT = Font(color="1B5E20", bold=True)
_NEG_FONT = Font(color="B71C1C", bold=True)
_BOLD = Font(bold=True)
_WHITE_BOLD = Font(color="FFFFFF", bold=True)
_CENTER = Alignment(horizontal="center")
_RIGHT = Alignment(horizontal="right")


def _sign_font(value: float) -> Font:
    return _POS_FONT if value >= 0 else _NEG_FONT


def export_xlsx(line_items: list[LineItem], year_label: str, path: str | Path):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Budget {year_label}"

    headers = ["Line item", "Type"] + MONTHS + ["Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _WHITE_BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER

    row = 2
    for li in line_items:
        fill = _CREDIT_FILL if li.category == "Credit" else _DEBIT_FILL
        ws.cell(row=row, column=1, value=li.label).fill = fill
        ws.cell(row=row, column=2, value=li.category).fill = fill
        for m in range(1, 13):
            cell = ws.cell(row=row, column=2 + m, value=li.amount(m))
            cell.number_format = '#,##0.00'
            cell.fill = fill
            cell.alignment = _RIGHT
        total = li.total()
        t_cell = ws.cell(row=row, column=15, value=total)
        t_cell.number_format = '#,##0.00'
        t_cell.fill = fill
        t_cell.font = _sign_font(total)
        t_cell.alignment = _RIGHT
        row += 1

    # separator
    row += 1

    credits = [li for li in line_items if li.category == "Credit"]
    debits = [li for li in line_items if li.category == "Debit"]

    for label_text, items in [("Total Credits", credits), ("Total Debits", debits)]:
        cell_label = ws.cell(row=row, column=1, value=label_text)
        cell_label.font = _BOLD
        cell_label.fill = _SUMMARY_FILL
        ws.cell(row=row, column=2).fill = _SUMMARY_FILL
        year_total = 0.0
        for m in range(1, 13):
            val = sum(li.amount(m) for li in items)
            year_total += val
            cell = ws.cell(row=row, column=2 + m, value=val)
            cell.number_format = '#,##0.00'
            cell.fill = _SUMMARY_FILL
            cell.font = _sign_font(val)
            cell.alignment = _RIGHT
        t_cell = ws.cell(row=row, column=15, value=year_total)
        t_cell.number_format = '#,##0.00'
        t_cell.fill = _SUMMARY_FILL
        t_cell.font = _sign_font(year_total)
        t_cell.alignment = _RIGHT
        row += 1

    net_row = row
    ws.cell(row=net_row, column=1, value="Net Balance").font = _BOLD
    ws.cell(row=net_row, column=1).fill = _SUMMARY_FILL
    ws.cell(row=net_row, column=2).fill = _SUMMARY_FILL
    net_year = 0.0
    for m in range(1, 13):
        tc = sum(li.amount(m) for li in credits)
        td = sum(li.amount(m) for li in debits)
        net = tc - td
        net_year += net
        cell = ws.cell(row=net_row, column=2 + m, value=net)
        cell.number_format = '#,##0.00'
        cell.fill = _SUMMARY_FILL
        cell.font = _sign_font(net)
        cell.alignment = _RIGHT
    t_cell = ws.cell(row=net_row, column=15, value=net_year)
    t_cell.number_format = '#,##0.00'
    t_cell.fill = _SUMMARY_FILL
    t_cell.font = _sign_font(net_year)
    t_cell.alignment = _RIGHT

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    for i in range(3, 16):
        ws.column_dimensions[get_column_letter(i)].width = 12

    wb.save(path)


def export_csv(line_items: list[LineItem], path: str | Path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Line item", "Type"] + MONTHS + ["Total"])
        for li in line_items:
            row = [li.label, li.category]
            row += [f"{li.amount(m):.2f}" for m in range(1, 13)]
            row.append(f"{li.total():.2f}")
            writer.writerow(row)

        writer.writerow([])
        credits = [li for li in line_items if li.category == "Credit"]
        debits = [li for li in line_items if li.category == "Debit"]
        for label_text, items in [("Total Credits", credits), ("Total Debits", debits)]:
            row = [label_text, ""]
            row += [f"{sum(li.amount(m) for li in items):.2f}" for m in range(1, 13)]
            row.append(f"{sum(li.total() for li in items):.2f}")
            writer.writerow(row)

        net_row = ["Net Balance", ""]
        for m in range(1, 13):
            tc = sum(li.amount(m) for li in credits)
            td = sum(li.amount(m) for li in debits)
            net_row.append(f"{tc - td:.2f}")
        net_row.append(f"{sum(li.total() for li in credits) - sum(li.total() for li in debits):.2f}")
        writer.writerow(net_row)
