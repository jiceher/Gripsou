"""Import wizard backend — reads xlsx, xlsb, csv into a neutral preview structure."""

from __future__ import annotations

import csv
from pathlib import Path


def read_file(path: str | Path) -> list[list[str]]:
    """Return raw rows as list[list[str]] from xlsx, xlsb or csv."""
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix == ".csv":
        return _read_csv(p)
    elif suffix == ".xlsx":
        return _read_xlsx(p)
    elif suffix == ".xlsb":
        return _read_xlsb(p)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")


def _read_csv(path: Path) -> list[list[str]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return [row for row in csv.reader(f)]


def _read_xlsx(path: Path) -> list[list[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    result = []
    for row in ws.iter_rows(values_only=True):
        result.append([str(c) if c is not None else "" for c in row])
    wb.close()
    return result


def _read_xlsb(path: Path) -> list[list[str]]:
    from pyxlsb import open_workbook
    result = []
    with open_workbook(str(path)) as wb:
        with wb.get_sheet(1) as ws:
            for row in ws.rows():
                result.append([str(r.v) if r.v is not None else "" for r in row])
    return result


def parse_import(
    raw_rows: list[list[str]],
    header_row: int,
    label_col: int,
    tag_col: int | None,
    month_cols: list[int],
    default_category: str,
) -> list[dict]:
    """
    Convert raw rows into a list of dicts ready for DB insertion.
    Each dict: {label, category, monthly_values: {1..12: float}}
    """
    results = []
    for i, row in enumerate(raw_rows):
        if i <= header_row:
            continue
        if label_col >= len(row) or not row[label_col].strip():
            continue
        label = row[label_col].strip()
        category = default_category
        if tag_col is not None and tag_col < len(row):
            raw_tag = row[tag_col].strip().lower()
            if raw_tag in ("credit", "income", "+"):
                category = "Credit"
            elif raw_tag in ("debit", "expense", "depense", "dépense", "-"):
                category = "Debit"
        monthly = {}
        for m_idx, col in enumerate(month_cols, 1):
            if col < len(row):
                try:
                    monthly[m_idx] = float(row[col].replace(",", ".").replace(" ", ""))
                except ValueError:
                    monthly[m_idx] = 0.0
            else:
                monthly[m_idx] = 0.0
        results.append({"label": label, "category": category, "monthly_values": monthly})
    return results
